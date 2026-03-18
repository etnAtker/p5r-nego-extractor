"""High-level orchestration for extracting negotiation data."""

from __future__ import annotations

import logging
import re
from collections import OrderedDict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from flow_parser import FlowParser
from models import FlatRow, MsgIndex, ReactionRule, TextEntry
from msg_parser import MESSAGE_BLOCK_RE, parse_msg_resources

LOGGER = logging.getLogger(__name__)


class NegotiationPipeline:
    def __init__(
        self,
        input_dir: Path,
        output_dir: Path,
        scripts: Optional[Sequence[str]] = None,
        patch_output_dir: Optional[Path] = None,
    ):
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.scripts = list(scripts) if scripts else self._discover_scripts()
        self.patch_output_dir = patch_output_dir

    def _discover_scripts(self) -> List[str]:
        scripts = []
        for path in sorted(self.input_dir.glob("TALK_*.BF.msg")):
            scripts.append(path.name.replace(".BF.msg", ""))
        return scripts

    def run(self) -> None:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        if self.patch_output_dir:
            self.patch_output_dir.mkdir(parents=True, exist_ok=True)

        all_text_entries: List[TextEntry] = []
        all_rules: List[ReactionRule] = []
        all_flat_rows: List[FlatRow] = []

        for script in self.scripts:
            LOGGER.info("Processing talk script %s", script)
            files = self._resolve_files(script)
            msg_index = parse_msg_resources(script, files.msg_path, files.header_path)
            flow_parser = FlowParser(script, files.flow_path, msg_index)
            rules = flow_parser.extract_rules()
            flat_rows = self._build_flat_rows(msg_index, rules)

            if self.patch_output_dir:
                self._write_compile_bundle(files, msg_index, rules)

            all_text_entries.extend(msg_index.entries)
            all_rules.extend(rules)
            all_flat_rows.extend(flat_rows)

        result_rows = self._build_result_rows(all_flat_rows)
        workbook_path = self.output_dir / "talk_negotiation_tables.xlsx"
        sheets = [
            ("text_index", all_text_entries, TEXT_FIELDS, TEXT_HEADERS_CN),
            ("rule_table", all_rules, RULE_FIELDS, RULE_HEADERS_CN),
            ("flat", all_flat_rows, FLAT_FIELDS, FLAT_HEADERS_CN),
        ]
        self._write_workbook(workbook_path, sheets, result_rows)
        self._write_result_text(self.output_dir / "talk_negotiation_summary.txt", result_rows)

    def _build_flat_rows(self, msg_index: MsgIndex, rules: Iterable[ReactionRule]) -> List[FlatRow]:
        rows: List[FlatRow] = []

        for rule in rules:
            if rule.reaction_msg_id is None or rule.question_msg_id is None:
                continue

            question_entry = msg_index.id_to_entry.get(rule.question_msg_id)
            reaction_entry = msg_index.id_to_entry.get(rule.reaction_msg_id)
            selection_entry = msg_index.selection_entries.get((rule.selection_msg_id, rule.choice_index))

            if not (question_entry and reaction_entry and selection_entry):
                continue

            row = FlatRow(
                talk_script=rule.talk_script,
                question_group=rule.question_group,
                question_no=rule.question_no,
                question_msg_id=rule.question_msg_id,
                question_msg_key=rule.question_msg_key,
                question_text_raw=question_entry.text_raw,
                question_text_clean=question_entry.text_clean,
                choice_index=rule.choice_index,
                choice_text_raw=selection_entry.text_raw,
                choice_text_clean=selection_entry.text_clean,
                personality_value=rule.personality_value,
                personality_label=rule.personality_label,
                reaction_grade_raw=rule.reaction_grade_raw,
                reaction_grade_label=rule.reaction_grade_label,
                reaction_msg_id=rule.reaction_msg_id,
                reaction_msg_key=rule.reaction_msg_key,
                reaction_text_raw=reaction_entry.text_raw,
                reaction_text_clean=reaction_entry.text_clean,
            )
            rows.append(row)

        return rows

    def _write_workbook(
        self,
        path: Path,
        sheets: List[Tuple[str, Iterable, List[str], List[str]]],
        result_rows: List[Tuple[dict, List[dict]]],
    ) -> None:
        wb = Workbook()
        # remove default sheet by reusing for first data set
        wb.remove(wb.active)

        for name, rows, fieldnames, headers in sheets:
            ws = wb.create_sheet(title=name)
            ws.append(headers)

            for row in rows:
                row_dict = asdict(row)
                ws.append([row_dict.get(field, "") for field in fieldnames])

        self._add_result_sheet(wb, result_rows)
        wb.save(path)
        LOGGER.info("Wrote %s", path)

    def _write_compile_bundle(
        self, files: "ScriptFiles", msg_index: MsgIndex, rules: Iterable[ReactionRule]
    ) -> None:
        if self.patch_output_dir is None:
            return

        hint_map = self._build_selection_hint_map(msg_index, rules)
        patched_msg = self._build_patched_msg_content(
            files.msg_path.read_text(encoding="utf-8", errors="ignore"),
            hint_map,
        )
        sanitized_msg = self._sanitize_msg_file_for_compile(patched_msg)

        output_msg = self.patch_output_dir / files.msg_path.name
        output_header = self.patch_output_dir / files.header_path.name
        output_flow = self.patch_output_dir / files.flow_path.name
        stale_minimal_patch = self.patch_output_dir / f"{files.msg_path.name.replace('.BF.msg', '')}.msg"

        if stale_minimal_patch.exists():
            stale_minimal_patch.unlink()

        output_msg.write_text(sanitized_msg, encoding="utf-8")
        output_header.write_text(files.header_path.read_text(encoding="utf-8", errors="ignore"), encoding="utf-8")
        output_flow.write_text(files.flow_path.read_text(encoding="utf-8", errors="ignore"), encoding="utf-8")
        LOGGER.info("Wrote compile bundle for %s to %s", files.msg_path.stem, self.patch_output_dir)

    def _add_result_sheet(self, wb: Workbook, result_rows: List[Tuple[dict, List[dict]]]) -> None:
        ws = wb.create_sheet(index=0, title="result")
        ws.append(RESULT_HEADERS)
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        header_font = Font(bold=True)
        for col_idx in range(1, len(RESULT_HEADERS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        ws.freeze_panes = "A2"
        current_row = 2

        script_ranges: List[Tuple[int, int]] = []
        question_ranges: List[Tuple[int, int]] = []
        current_script = None
        current_script_start = None
        current_script_end = None

        for meta, options in result_rows:
            if not options:
                continue

            for idx, option in enumerate(options):
                ws.append(
                    [
                        meta["script"] if idx == 0 else "",
                        meta["text"] if idx == 0 else "",
                        option["choice_text"],
                        option["开朗"],
                        option["懦弱"],
                        option["性急"],
                        option["阴沉"],
                    ]
                )

            row_span = len(options)
            script_value = meta["script"]

            if row_span > 0:
                question_ranges.append((current_row, current_row + row_span - 1))

            if current_script != script_value:
                if current_script is not None and current_script_start is not None and current_script_end is not None:
                    script_ranges.append((current_script_start, current_script_end))
                current_script = script_value
                current_script_start = current_row

            current_script_end = current_row + row_span - 1

            if row_span > 1:
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + row_span - 1, end_column=2)

            current_row += row_span

        if current_script is not None and current_script_start is not None and current_script_end is not None:
            script_ranges.append((current_script_start, current_script_end))

        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")

        question_starts = {start for start, _ in question_ranges}
        script_starts = {start for start, _ in script_ranges}

        for row in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=7):
            for cell in row:
                row_idx = cell.row
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                if cell.column in (1, 2):
                    cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
                    continue

                if row_idx == 1:
                    top = thin
                    bottom = thick
                else:
                    top = thick if row_idx in question_starts or row_idx in script_starts else thin
                    bottom = thin

                cell.border = Border(left=thin, right=thin, top=top, bottom=bottom)

        col_widths = {1: 15, 2: 115, 3: 18, 4: 10, 5: 10, 6: 10, 7: 10}
        for idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width

        for start, end in script_ranges:
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    def _build_result_rows(self, flat_rows: Iterable[FlatRow]) -> List[Tuple[dict, List[dict]]]:
        questions: "OrderedDict[Tuple[str, str, str], dict]" = OrderedDict()

        for row in flat_rows:
            key = (
                row.talk_script,
                row.question_no or "",
                row.question_text_clean or row.question_text_raw or "",
            )

            if key not in questions:
                questions[key] = {
                    "meta": {"script": key[0], "number": key[1], "text": key[2]},
                    "options": {idx: self._new_option_entry() for idx in range(1, RESULT_OPTION_COUNT + 1)},
                }

            choice_index = row.choice_index
            if choice_index is None or choice_index < 1 or choice_index > RESULT_OPTION_COUNT:
                continue

            option_entry = questions[key]["options"][choice_index]
            if not option_entry["choice_text"]:
                option_entry["choice_text"] = row.choice_text_clean or row.choice_text_raw or ""

            personality_label = row.personality_label or ""
            reaction_grade_label = row.reaction_grade_label or ""

            if personality_label in option_entry and reaction_grade_label in REACTION_SYMBOLS:
                option_entry[personality_label] = REACTION_SYMBOLS[reaction_grade_label]

        result_data: List[Tuple[dict, List[dict]]] = []

        for question in questions.values():
            option_rows: List[dict] = []
            for idx in range(1, RESULT_OPTION_COUNT + 1):
                entry = question["options"][idx]
                if not entry["choice_text"]:
                    continue
                option_rows.append(
                    {
                        "choice_text": entry["choice_text"],
                        "开朗": entry["开朗"],
                        "懦弱": entry["懦弱"],
                        "性急": entry["性急"],
                        "阴沉": entry["阴沉"],
                        "喜欢": self._format_personalities_by_symbol(entry, "√"),
                        "一般": self._format_personalities_by_symbol(entry, "-"),
                        "反感": self._format_personalities_by_symbol(entry, "X"),
                    }
                )

            if option_rows:
                result_data.append((question["meta"], option_rows))

        return result_data

    def _write_result_text(self, path: Path, result_rows: List[Tuple[dict, List[dict]]]) -> None:
        def _fmt(text: Optional[str]) -> str:
            return text.replace("\n", " ").strip() if text else ""

        with path.open("w", encoding="utf-8") as handle:
            first = True
            for meta, options in result_rows:
                if not options:
                    continue

                if not first:
                    handle.write("\n------\n\n")
                first = False

                question_label = _fmt(meta["text"])
                handle.write(f"{question_label}\n")

                for option in options:
                    choice_label = _fmt(option["choice_text"])
                    like = option["喜欢"] or ""
                    normal = option["一般"] or ""
                    dislike = option["反感"] or ""
                    handle.write(
                        f"> {choice_label} >> [{like}] - 喜欢 / [{normal}] - 一般 / [{dislike}] - 讨厌\n"
                    )

        LOGGER.info("Wrote %s", path)

    def _build_selection_hint_map(
        self, msg_index: MsgIndex, rules: Iterable[ReactionRule]
    ) -> Dict[str, Dict[int, str]]:
        selection_data: "OrderedDict[str, Dict[int, Dict[str, str]]]" = OrderedDict()

        for selection_id, choice_count in sorted(msg_index.selection_choice_counts.items()):
            selection_key = msg_index.id_to_key.get(selection_id)
            if not selection_key:
                continue
            selection_data[selection_key] = {
                choice_index: {label: "" for label in PERSONALITY_ORDER}
                for choice_index in range(1, choice_count + 1)
            }

        for rule in rules:
            selection_key = rule.selection_msg_key or msg_index.id_to_key.get(rule.selection_msg_id)
            if not selection_key:
                continue

            choice_entry = selection_data.setdefault(selection_key, {})
            personalities = choice_entry.setdefault(
                rule.choice_index,
                {label: "" for label in PERSONALITY_ORDER},
            )

            if rule.personality_label in personalities and rule.reaction_grade_label in REACTION_SYMBOLS:
                personalities[rule.personality_label] = REACTION_SYMBOLS[rule.reaction_grade_label]

        hint_map: Dict[str, Dict[int, str]] = {}
        for selection_key, choices in selection_data.items():
            hint_map[selection_key] = {}
            for choice_index, personalities in choices.items():
                likes = "".join(
                    PERSONALITY_ABBREVIATIONS[label]
                    for label in PERSONALITY_ORDER
                    if personalities.get(label) == REACTION_SYMBOLS["喜欢"]
                )
                normals = "".join(
                    PERSONALITY_ABBREVIATIONS[label]
                    for label in PERSONALITY_ORDER
                    if personalities.get(label) == REACTION_SYMBOLS["一般"]
                )
                dislikes = "".join(
                    PERSONALITY_ABBREVIATIONS[label]
                    for label in PERSONALITY_ORDER
                    if personalities.get(label) == REACTION_SYMBOLS["反感"]
                )
                segments: List[str] = []
                if likes:
                    segments.append(f"[clr {RESULT_HINT_COLORS['喜欢']}]{likes}")
                if normals:
                    segments.append(f"[clr {RESULT_HINT_COLORS['一般']}]{normals}")
                if dislikes:
                    segments.append(f"[clr {RESULT_HINT_COLORS['反感']}]{dislikes}")

                if segments:
                    hint_map[selection_key][choice_index] = f"({''.join(segments)}[clr {RESULT_HINT_RESET_COLOR}])"
                else:
                    hint_map[selection_key][choice_index] = ""

        return hint_map

    def _build_patched_msg_content(self, content: str, hint_map: Dict[str, Dict[int, str]]) -> str:
        lines = content.splitlines(keepends=True)
        output_lines: List[str] = []
        current_header: Optional[str] = None
        current_type: Optional[str] = None
        current_key: Optional[str] = None
        buffer: List[str] = []

        def _flush_block() -> None:
            nonlocal current_header, current_type, current_key, buffer
            if current_header is None:
                return

            block_text = "".join(buffer)
            if current_type == "sel" and current_key in hint_map:
                block_text = self._patch_selection_block(block_text, hint_map[current_key])

            output_lines.append(current_header)
            output_lines.append(block_text)

            current_header = None
            current_type = None
            current_key = None
            buffer = []

        for line in lines:
            match = MESSAGE_BLOCK_RE.match(line)
            if match:
                _flush_block()
                current_header = line
                current_type = match.group(1)
                current_key = match.group(2)
                buffer = []
            else:
                buffer.append(line)

        _flush_block()
        return "".join(output_lines)

    @staticmethod
    def _patch_selection_block(block_text: str, choice_hints: Dict[int, str]) -> str:
        choice_counter = 0

        def _replace(match) -> str:
            nonlocal choice_counter
            choice_counter += 1
            hint = choice_hints.get(choice_counter)
            if not hint:
                return match.group(0)

            option_prefix = match.group(1)
            option_body = match.group(2)
            option_end = match.group(3)
            trimmed_body = option_body.rstrip()
            trailing = option_body[len(trimmed_body) :]
            return f"{option_prefix}{trimmed_body}{hint}{trailing}{option_end}"

        return SELECTION_OPTION_RE.sub(_replace, block_text)

    @staticmethod
    def _sanitize_msg_text_for_compile(content: str) -> str:
        parts = TAG_OR_ESCAPED_BRACKET_RE.split(content)
        sanitized: List[str] = []

        for part in parts:
            if not part:
                continue

            if TAG_OR_ESCAPED_BRACKET_RE.fullmatch(part):
                sanitized.append(part)
                continue

            for ch in part:
                sanitized.append(_normalize_ascii_for_p5r_chs(ch))

        return "".join(sanitized)

    def _sanitize_msg_file_for_compile(self, content: str) -> str:
        sanitized_lines: List[str] = []
        for line in content.splitlines(keepends=True):
            if line.endswith("\r\n"):
                line_ending = "\r\n"
                line_body = line[:-2]
            elif line.endswith("\n"):
                line_ending = "\n"
                line_body = line[:-1]
            else:
                line_ending = ""
                line_body = line

            match = MESSAGE_BLOCK_RE.match(line_body)
            if match:
                sanitized_lines.append(self._sanitize_header_line(line_body) + line_ending)
            else:
                sanitized_lines.append(self._sanitize_msg_text_for_compile(line_body) + line_ending)

        return "".join(sanitized_lines)

    def _sanitize_header_line(self, line: str) -> str:
        speaker_match = MESSAGE_SPEAKER_HEADER_RE.match(line)
        if not speaker_match:
            return line

        prefix = speaker_match.group(1)
        speaker = speaker_match.group(2)
        suffix = speaker_match.group(3)
        sanitized_speaker = self._sanitize_msg_text_for_compile(speaker)
        return f"{prefix}{sanitized_speaker}{suffix}"

    @staticmethod
    def _new_option_entry() -> dict:
        return {
            "choice_text": "",
            "开朗": "",
            "懦弱": "",
            "性急": "",
            "阴沉": "",
        }

    @staticmethod
    def _format_personalities_by_symbol(option_entry: dict, symbol: str) -> str:
        labels = [label for label in PERSONALITY_ORDER if option_entry.get(label) == symbol]
        return "、".join(labels)

    def _resolve_files(self, script: str) -> "ScriptFiles":
        msg_path = self.input_dir / f"{script}.BF.msg"
        header_path = self.input_dir / f"{script}.BF.msg.h"
        flow_path = self.input_dir / f"{script}.BF.flow"

        for path in (msg_path, header_path, flow_path):
            if not path.exists():
                raise FileNotFoundError(f"Required file not found: {path}")

        return ScriptFiles(msg_path=msg_path, header_path=header_path, flow_path=flow_path)


TEXT_FIELDS = [
    "talk_script",
    "msg_id",
    "msg_key",
    "msg_type",
    "source_file",
    "question_group",
    "question_no",
    "choice_index",
    "text_raw",
    "text_clean",
]

RULE_FIELDS = [
    "talk_script",
    "variant",
    "question_group",
    "question_no",
    "question_msg_id",
    "question_msg_key",
    "selection_msg_id",
    "selection_msg_key",
    "choice_raw",
    "choice_index",
    "personality_value",
    "personality_label",
    "reaction_grade_raw",
    "reaction_grade_label",
    "reaction_offset",
    "reaction_msg_id",
    "reaction_msg_key",
]

FLAT_FIELDS = [
    "talk_script",
    "question_group",
    "question_no",
    "question_msg_id",
    "question_msg_key",
    "question_text_raw",
    "question_text_clean",
    "choice_index",
    "choice_text_raw",
    "choice_text_clean",
    "personality_value",
    "personality_label",
    "reaction_grade_raw",
    "reaction_grade_label",
    "reaction_msg_id",
    "reaction_msg_key",
    "reaction_text_raw",
    "reaction_text_clean",
]

TEXT_HEADERS_CN = [
    "脚本",
    "消息ID",
    "消息键",
    "类型",
    "源文件",
    "题组",
    "题号",
    "选项序号",
    "原始文本",
    "清洗文本",
]

RULE_HEADERS_CN = [
    "脚本",
    "题组",
    "题号",
    "题目ID",
    "题目键",
    "选项ID",
    "选项键",
    "选项序号(原始)",
    "选项序号",
    "人格值",
    "人格标签",
    "反应等级值",
    "反应等级标签",
    "反应偏移",
    "反应ID",
    "反应键",
]

FLAT_HEADERS_CN = [
    "脚本",
    "题组",
    "题号",
    "题目ID",
    "题目键",
    "题目原文",
    "题目文本",
    "选项序号",
    "选项原文",
    "选项文本",
    "人格值",
    "人格标签",
    "反应等级值",
    "反应等级标签",
    "反应ID",
    "反应键",
    "反应原文",
    "反应文本",
]

RESULT_HEADERS = ["脚本", "题目", "选项", "开朗", "懦弱", "性急", "阴沉"]
RESULT_OPTION_COUNT = 4
PERSONALITY_ORDER = ["开朗", "懦弱", "性急", "阴沉"]
PERSONALITY_ABBREVIATIONS = {"开朗": "开", "懦弱": "懦", "性急": "急", "阴沉": "阴"}
REACTION_SYMBOLS = {"喜欢": "√", "一般": "-", "反感": "X"}
RESULT_HINT_COLORS = {"喜欢": 4, "一般": 26, "反感": 10}
RESULT_HINT_RESET_COLOR = 27
SELECTION_OPTION_RE = re.compile(r"(\[s[^\]]*\])(.*?)(\[e\])", flags=re.DOTALL | re.IGNORECASE)
TAG_OR_ESCAPED_BRACKET_RE = re.compile(r"(\\\[|\\\]|\[[^\]]+\])")
MESSAGE_SPEAKER_HEADER_RE = re.compile(r"^(\[msg\s+[^\]\s]+\s+\[)(.*?)(\]\])$")


def _normalize_ascii_for_p5r_chs(ch: str) -> str:
    if ch == " ":
        return ""

    code = ord(ch)
    if 0x21 <= code <= 0x7E:
        return chr(code + 0xFEE0)

    return ch


@dataclass
class ScriptFiles:
    msg_path: Path
    header_path: Path
    flow_path: Path
